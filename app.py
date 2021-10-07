import openpyxl as xl

def process_workbook(filename):
    wb = xl.load_workbook(filename)  # importing file,use Tab for crcting
    sheet = wb["Sheet1"]  # sheet name
    cell = sheet["a1"]  # index number
    cell = sheet.cell(1, 1)  # rows and colums
    # print(sheet.max_row)  # for finding how many rows
    # print(cell.value)
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price